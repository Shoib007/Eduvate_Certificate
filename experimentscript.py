import openpyxl, os, argparse
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4, landscape
from datetime import datetime

# Read the cert_type from the keyword argument if not exist return error
parser = argparse.ArgumentParser()
parser.add_argument("--cert_type", type=str, help="The type of certificate to generate", required=True, choices=["CSR", "ROBOFEST", "TEACHERS"])
args = parser.parse_args()


# get working directory
cwd = os.getcwd()

# Read all the files in the cwd/data folder
data_folder = os.path.join(cwd, "data")
output_folder = os.path.join(cwd, "certificates")

# Check if the output_folder exists or not
if not os.path.exists(output_folder):
    os.makedirs(output_folder)


data_files = os.listdir(data_folder)


# Get current date in DD/MM/YYYY format
today = datetime.today().strftime("%d/%m/%Y")


# Get template path
template_path = os.path.join(cwd, "CSR_template.jpg")


# Titles of Certificate
cs_title_experties = {
    "GRADE 1": ["Emerging Coder", "Block Based Coding"],
    "GRADE 2": ["Junior Coder", "Block Based Coding"],
    "GRADE 3": ["Pro Coder", "Block Based Coding"],
    "GRADE 4": ["Emerging App Develoer", "App Lab"],
    "GRADE 5": ["Pro App Develoer", "MIT App Inventor"],
    "GRADE 6": ["Junior Web Develoer", "Web Development"],
    "GRADE 7": ["Pro Web Develoer", "Web Development"],
    "GRADE 8": ["Junior Pythonista", "Python 3"],
    "GRADE 9": ["Emerging AI Developer", "Artificial Intelligence"],
}


def generate_cs_robotics_certificate(schoolName, grade, student_name, output_folder, template_path):
    c = canvas.Canvas(f"{output_folder}/{grade}_{student_name}.pdf", pagesize=landscape(A4))
    c.drawImage(template_path, 0, 0, landscape(A4)[0], landscape(A4)[1])

    # Customize certificate with data
    c.setFont("Helvetica", 24)
    c.setFillColorRGB(0, 0, 0)
    c.drawCentredString(420, 350, student_name)  # Adjust coordinates
    c.setFont("Helvetica", 20)
    c.drawCentredString(280, 310,  grade.split()[1])  # Adjust coordinates
    c.drawCentredString(600, 310, schoolName.capitalize())  
    c.drawCentredString(550, 260, cs_title_experties[grade][0])
    c.drawCentredString(580, 175, cs_title_experties[grade][1])
    c.setFont("Helvetica", 12)
    c.drawCentredString(180, 110, today)

    c.save()


def generate_robofest_certificate(schoolName, grade, student_name, output_folder, template_path):
    c = canvas.Canvas(f"{output_folder}/{grade}_{student_name}.pdf", pagesize=landscape(A4))
    c.drawImage(template_path, 0, 0, landscape(A4)[0], landscape(A4)[1])

    # Customize certificate with data
    c.setFont("Helvetica", 24)
    c.setFillColorRGB(0, 0, 0)
    c.drawCentredString(420, 350, student_name)  # Adjust coordinates
    c.setFont("Helvetica", 20)
    c.drawCentredString(280, 310, grade.split()[0])  # Adjust coordinates
    c.drawCentredString(600, 310, schoolName)  
    c.drawCentredString(550, 260, cs_title_experties[grade][0])
    c.drawCentredString(580, 175, cs_title_experties[grade][1])
    c.setFont("Helvetica", 12)
    c.drawCentredString(180, 110, today)

    c.save()



# Main Logic for generating certificates
for school_file in data_files:
    schoolName = school_file.split(".")[0].upper()
    print(schoolName)

    # Load the school sheet
    ws = openpyxl.load_workbook(os.path.join(data_folder, school_file))

    # Get the list of sheet names within that school sheet
    sheetNames = ws.sheetnames

    for grade in sheetNames:
        sheet = ws[grade]

        # Creating folder for each school and grade
        folder = os.path.join(cwd, output_folder, schoolName , grade)
        if not os.path.exists(folder):
            os.makedirs(folder)

        for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header
            student_name = row[0]
            grade = grade.upper()

            if args.cert_type == "CSR":
                generate_cs_robotics_certificate(schoolName, grade, student_name, folder, template_path)
                print(f"{schoolName}-{grade}:{student_name}")
            elif args.cert_type == "ROBOFEST":
                generate_robofest_certificate(schoolName, grade, student_name, folder, template_path)
                print(f"{schoolName}-{grade}:{student_name}")
