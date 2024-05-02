from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import openpyxl, os, argparse
from robofestCertificate import generate_robofest_certificate
from CSR_Certificate import generate_cs_robotics_certificate
from teachersCertificate import generate_teachers_certificate
from datetime import datetime


# Read the cert_type from the keyword argument if not exist return error
parser = argparse.ArgumentParser()
parser.add_argument("--cert_type", type=str, help="The type of certificate to generate", required=True, choices=["CS", "ROBOTICS", "ROBOFEST", "TEACHER", "CSR_COMBO"])
parser.add_argument("--subject", type=str, help="Subject for Teachers certificate", required=False, choices=["CS", "ROBOTICS", "CSR"], default="CS")
args = parser.parse_args()

# get working directory
cwd = os.getcwd()

# Read all the files in the cwd/data folder
data_folder = os.path.join(cwd, "data")
output_folder = os.path.join(cwd, "certificates")


# Check if the output_folder exists or not
if not os.path.exists(output_folder):
    os.makedirs(output_folder)


data_files = os.listdir(data_folder)



# Get template path
CSR_template = os.path.join(cwd, "template","CSR_template.png")
Robofest_template = os.path.join(cwd,"template", "Robofest_template.png")
teacher_template = os.path.join(cwd, "template", "Teachers_template.png")

# Register fonts here
pdfmetrics.registerFont(TTFont("Roboto-Light-Italic", "Roboto-LightItalic.ttf"))
pdfmetrics.registerFont(TTFont("Roboto-Regular", "Roboto-Regular.ttf"))
pdfmetrics.registerFont(TTFont("Dancing-Font", "DancingScript-Regular.ttf"))
pdfmetrics.registerFont(TTFont("Roboto-Italic", "Roboto-Italic.ttf"))



# Main Logic for generating certificates
for data in data_files:
    schoolName = data.split(".")[0]
    print(f"Creating certificate for {schoolName} -> {args.cert_type} ...")

    # Load the school sheet
    ws = openpyxl.load_workbook(os.path.join(data_folder, data), data_only=True)

    # Get the list of sheet names within that school sheet
    sheetNames = ws.sheetnames
    for grade in sheetNames:
        sheet = ws[grade]
        folder = os.path.join(cwd, output_folder, schoolName , args.cert_type, grade)

        for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header
            if args.cert_type == "CS" or args.cert_type == "ROBOTICS" or args.cert_type == "CSR_COMBO":
                # Creating folder for each school, Subject and grade
                if not os.path.exists(folder):
                    os.makedirs(folder)

                if row[0] is None:
                    break
                student_name = row[0].title()
                grade = grade.title()

                generate_cs_robotics_certificate(schoolName, grade, student_name, folder, CSR_template, args.cert_type)
            

            
            elif args.cert_type == "ROBOFEST":
                if row[0] is None:
                    break
                
                # Convert the date string to datetime object by removing 00:00:00
                date = datetime.strptime(str(row[0]), '%Y-%m-%d %H:%M:%S').strftime('%d-%m-%Y')
                student_name = row[1].title()
                grade = str(row[2])
                projectName = row[4].title()

                # Creating folder for each school, Subject and grade
                folder = os.path.join(cwd, output_folder, schoolName , args.cert_type)
                if not os.path.exists(folder):
                    os.makedirs(folder)

                generate_robofest_certificate(date, student_name, grade, schoolName, projectName, output_folder=folder, template_path=Robofest_template)
            elif args.cert_type == "TEACHER":
                if row[0] is None:
                    break

                teacher_name = row[0].title()
                school_name = row[1].title()
                total_hr_of_training = row[2]
                grades = row[3]

                folder = os.path.join(output_folder, school_name, args.subject )
                if not os.path.exists(folder):
                    os.makedirs(folder)

                if args.subject == "CS": generate_teachers_certificate(teacher_name, grades, school_name, total_hr_of_training, folder, teacher_template, subject="Computer Science")

                elif args.subject == "ROBOTICS": generate_teachers_certificate(teacher_name, grades, school_name, total_hr_of_training, folder, teacher_template, subject="Robotics")

                elif args.subject == "CSR":
                    generate_teachers_certificate(teacher_name, grades, school_name, total_hr_of_training, folder, teacher_template, subject="Computer Science")

                    generate_teachers_certificate(teacher_name, grades, school_name, total_hr_of_training, folder, teacher_template, subject="Robotics")
                
                print(f"Certificate for {teacher_name} -> {args.cert_type} created successfully!")
                


    if not args.cert_type == "TEACHER":
        print(f"Certificate for {schoolName} -> {args.cert_type} created successfully!")

